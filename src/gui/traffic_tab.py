import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import sqlite3
import json
import os
from datetime import datetime, timedelta, time  # Додано timedelta
from math import radians, sin, cos, sqrt, atan2, degrees
from shapely.geometry import Point, shape
from typing import List, Dict, Optional, Tuple
from pathlib import Path
import logging
from fuzzywuzzy import fuzz
import re
from ..utils.config import Config
from ..core.data_processor import DataProcessor
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import numpy as np
import folium
from folium import plugins

class TrafficTab(ttk.Frame):
    def __init__(self, parent: ttk.Notebook, config: 'Config', data_processor: 'DataProcessor'):
        """
        Ініціалізація вкладки обробки трафіку.

        Args:
            parent: Батьківський віджет (ttk.Notebook)
            config: Конфігурація програми
            data_processor: Обробник даних
        """
        super().__init__(parent)  # Використовуємо parent замість notebook

        # Зберігаємо залежності
        self.config = config
        self.data_processor = data_processor

        # Ініціалізація змінних
        self.current_time = datetime.now()
        self.current_user = os.getenv('USERNAME', 'Unknown')
        self.traffic_files = []
        self.date_filter_file = None

        # Створюємо прогрес-бар
        self.progress_bar = ttk.Progressbar(self, mode='determinate')

        # Створюємо інтерфейс
        self._create_widgets()

    def _create_widgets(self):
        """Створення віджетів інтерфейсу."""
        # Створюємо головний контейнер з трьома колонками
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        main_frame.columnconfigure(0, weight=1)

        # Фрейм для кнопок (верхній ряд)
        buttons_frame = ttk.LabelFrame(main_frame, text="Операції з файлами")
        buttons_frame.grid(row=0, column=0, sticky='ew', padx=5, pady=5)

        # Контейнер для кнопок з можливістю переносу на новий рядок
        button_container = ttk.Frame(buttons_frame)
        button_container.pack(fill=tk.X, padx=5, pady=5)

        # Кнопки для роботи з файлами
        buttons = [
            ("Вибрати файли трафіку", self._select_traffic_files),
            ("Вибрати GeoJSON", self._select_geojson),
            ("Обробити", self._process_files),
            ("Об'єднати файли", lambda: self.data_processor.merge_traffic_files()),
            ("Знайти зустрічі", self.find_meetings),
            ("Аналіз активності", self._analyze_activity)
        ]

        for i, (text, command) in enumerate(buttons):
            ttk.Button(
                button_container,
                text=text,
                command=command
            ).grid(row=i // 3, column=i % 3, padx=5, pady=5, sticky='ew')

        # Мітка для відображення вибраних файлів
        self.files_label = ttk.Label(buttons_frame, text="Файли не вибрано")
        self.files_label.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)

        # Параметри обробки (середній ряд)
        params_frame = ttk.LabelFrame(main_frame, text="Параметри")
        params_frame.grid(row=1, column=0, sticky='ew', padx=5, pady=5)

        # Створюємо фрейми для кожної групи параметрів
        connection_frame = ttk.Frame(params_frame)
        connection_frame.pack(fill=tk.X, padx=5, pady=5)

        # Тип з'єднання
        ttk.Label(connection_frame, text="Тип з'єднання:").pack(side=tk.LEFT, padx=5)
        self.connection_type = tk.StringVar(value="Всі")
        connection_types = ["Всі", "вих", "вих СМС", "вх", "вх СМС", "переадр"]
        ttk.Combobox(
            connection_frame,
            textvariable=self.connection_type,
            values=connection_types,
            width=15
        ).pack(side=tk.LEFT, padx=5)

        # Часове вікно
        ttk.Label(connection_frame, text="Часове вікно (хв):").pack(side=tk.LEFT, padx=5)
        self.time_window = ttk.Entry(
            connection_frame,
            width=10
        )
        self.time_window.insert(0, "30")
        self.time_window.pack(side=tk.LEFT, padx=5)

        # Максимальна відстань
        ttk.Label(connection_frame, text="Макс. відстань (м):").pack(side=tk.LEFT, padx=5)
        self.max_distance = tk.StringVar(value="400")
        ttk.Entry(
            connection_frame,
            textvariable=self.max_distance,
            width=10
        ).pack(side=tk.LEFT, padx=5)

        # Фрейм для параметрів секторів
        sector_frame = ttk.Frame(params_frame)
        sector_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(sector_frame, text="Радіус сектору (м):").pack(side=tk.LEFT, padx=5)
        self.sector_radius = tk.StringVar(value="500")
        ttk.Entry(
            sector_frame,
            textvariable=self.sector_radius,
            width=10
        ).pack(side=tk.LEFT, padx=5)

        ttk.Label(sector_frame, text="Кут сектору (°):").pack(side=tk.LEFT, padx=5)
        self.sector_angle = tk.StringVar(value="120")
        ttk.Entry(
            sector_frame,
            textvariable=self.sector_angle,
            width=10
        ).pack(side=tk.LEFT, padx=5)

        # Фрейм для аналізу за конкретну дату
        date_frame = ttk.Frame(params_frame)
        date_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(date_frame, text="Дата (DD.MM.YYYY):").pack(side=tk.LEFT, padx=5)
        self.specific_date = tk.StringVar()
        ttk.Entry(
            date_frame,
            textvariable=self.specific_date,
            width=10
        ).pack(side=tk.LEFT, padx=5)

        # Фрейм фільтрації за датою
        # Фрейм фільтрації за датою
        filter_frame = ttk.LabelFrame(main_frame, text="Фільтрація за датою")
        filter_frame.grid(row=2, column=0, sticky='ew', padx=5, pady=5)

        # Додаємо часовий діапазон
        time_window_frame = ttk.Frame(filter_frame)
        time_window_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(time_window_frame, text="Часовий діапазон (±хв):").pack(side=tk.LEFT, padx=5)
        self.time_range = ttk.Entry(time_window_frame, width=10)
        self.time_range.insert(0, "10")  # За замовчуванням ±10 хвилин
        self.time_range.pack(side=tk.LEFT, padx=5)

        # Кнопки фільтрації (тільки один набір)
        filter_buttons = ttk.Frame(filter_frame)
        filter_buttons.pack(fill=tk.X, padx=5, pady=5)

        ttk.Button(
            filter_buttons,
            text="Вибрати файл з датами",
            command=self._select_date_filter_file
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            filter_buttons,
            text="Фільтрувати за датою",
            command=self._filter_by_date
        ).pack(side=tk.LEFT, padx=5)

        # Фрейм для логу операцій (нижній ряд)
        log_frame = ttk.LabelFrame(main_frame, text="Лог операцій")
        log_frame.grid(row=3, column=0, sticky='nsew', padx=5, pady=5)
        main_frame.rowconfigure(3, weight=1)

        # Текстове поле для логу
        self.log_text = tk.Text(log_frame, wrap=tk.WORD, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)

        # Скролбар для логу
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.configure(yscrollcommand=scrollbar.set)

        # Прогрес-бар (внизу вікна)
        self.progress_bar.pack(fill=tk.X, padx=5, pady=5)

        # Додаємо початковий запис в лог
        self.log_text.insert(
            tk.END,
            f"Current Date and Time (UTC - YYYY-MM-DD HH:MM:SS formatted): "
            f"{datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"Current User's Login: {self.current_user}\n"
            "Система готова до роботи\n\n"
        )
        self.log_text.see(tk.END)

    def _backup_database(self):
        """Експорт бази даних в Excel файл."""
        try:
            # Підключаємось до бази
            conn = sqlite3.connect("addresses.db")
            cursor = conn.cursor()

            # Отримуємо всі записи (змінюємо запит щоб отримати тільки існуючі колонки)
            cursor.execute(
                "SELECT address, latitude, longitude FROM addresses"
            )
            rows = cursor.fetchall()

            if not rows:
                messagebox.showwarning(
                    "Попередження",
                    "База даних порожня"
                )
                return

            # Створюємо DataFrame
            df = pd.DataFrame(
                rows,
                columns=['Адреса', 'Широта', 'Довгота']
            )

            # Вибираємо шлях для збереження
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel файли", "*.xlsx")],
                title="Зберегти базу даних як"
            )

            if not file_path:
                return

            # Зберігаємо в Excel
            df.to_excel(file_path, index=False)

            messagebox.showinfo(
                "Успіх",
                f"База даних експортована в {file_path}\n"
                f"Збережено {len(rows)} записів"
            )

            conn.close()

        except Exception as e:
            messagebox.showerror(
                "Помилка",
                f"Помилка експорту бази даних: {e}"
            )

    def _restore_database(self):
        """Відновлення бази даних з Excel файлу."""
        try:
            # Вибираємо файл для імпорту
            file_path = filedialog.askopenfilename(
                title="Виберіть файл для відновлення бази",
                filetypes=[("Excel файли", "*.xlsx")]
            )

            if not file_path:
                return

            # Імпортуємо дані
            conn = self.data_processor.import_addresses_to_sqlite(
                file_path=file_path
            )

            if conn:
                # Перевіряємо кількість записів
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM addresses")
                count = cursor.fetchone()[0]

                messagebox.showinfo(
                    "Успіх",
                    f"База даних відновлена з {file_path}\n"
                    f"Імпортовано {count} записів"
                )

                conn.close()

        except Exception as e:
            messagebox.showerror(
                "Помилка",
                f"Помилка відновлення бази даних: {e}"
            )

    def _recreate_database(self):
        """Очищення і перестворення бази даних."""
        try:
            # Видаляємо стару базу
            if os.path.exists("addresses.db"):
                os.remove("addresses.db")

            # Створюємо нову базу
            conn = sqlite3.connect("addresses.db")
            cursor = conn.cursor()

            # Створюємо таблицю
            cursor.execute('''
                CREATE TABLE addresses (
                    address TEXT PRIMARY KEY,
                    original_address TEXT,
                    latitude REAL,
                    longitude REAL
                )
            ''')

            # Створюємо індекс
            cursor.execute('CREATE INDEX idx_address ON addresses (address)')

            conn.commit()
            conn.close()

            self.log_text.insert(
                tk.END,
                "База даних успішно очищена і перестворена\n"
            )
            self.log_text.see(tk.END)

            messagebox.showinfo(
                "Успіх",
                "База даних успішно очищена і перестворена"
            )

        except Exception as e:
            error_msg = f"Помилка очищення бази даних: {e}"
            self.log_text.insert(tk.END, f"{error_msg}\n")
            self.log_text.see(tk.END)
            messagebox.showerror(
                "Помилка",
                error_msg
            )

    def _show_database_content(self):
        """Показати вміст бази даних."""
        try:
            conn = sqlite3.connect("addresses.db")
            cursor = conn.cursor()

            # Отримуємо всі записи
            cursor.execute("SELECT * FROM addresses")
            rows = cursor.fetchall()

            if not rows:
                messagebox.showinfo("База даних", "База даних порожня")
                return

            # Створюємо нове вікно
            window = tk.Toplevel(self)
            window.title("Вміст бази даних")
            window.geometry("800x600")

            # Створюємо текстове поле з прокруткою
            text = tk.Text(window, wrap=tk.WORD)
            text.pack(fill=tk.BOTH, expand=True)

            # Додаємо записи
            for row in rows:
                text.insert(tk.END, f"Адреса: {row[0]}\n")
                text.insert(tk.END, f"Оригінальна адреса: {row[1]}\n")
                text.insert(tk.END, f"Широта: {row[2]}\n")
                text.insert(tk.END, f"Довгота: {row[3]}\n")
                text.insert(tk.END, "-" * 50 + "\n")

            conn.close()

        except Exception as e:
            messagebox.showerror("Помилка", f"Помилка перегляду бази даних: {e}")

    def _load_base_stations(self):
        """Завантаження базових станцій з Excel файлу."""
        try:
            # Вибір файлу Excel
            file_path = filedialog.askopenfilename(
                title="Виберіть файл з базовими станціями",
                filetypes=[("Excel файли", "*.xlsx"), ("Всі файли", "*.*")]
            )

            if not file_path:
                return

            # Завантаження даних в базу
            conn = self.data_processor.import_addresses_to_sqlite(
                file_path=file_path
            )

            if conn:
                # Перевіряємо кількість записів
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM addresses")
                count = cursor.fetchone()[0]

                self.log_text.insert(
                    tk.END,
                    f"Завантажено базові станції з файлу: {file_path}\n"
                    f"Всього записів в базі: {count}\n"
                )
                self.log_text.see(tk.END)

                conn.close()

                # Показуємо повідомлення про успіх
                messagebox.showinfo(
                    "Успіх",
                    f"Базові станції успішно завантажено\n"
                    f"Всього записів в базі: {count}"
                )

        except Exception as e:
            error_msg = f"Помилка завантаження базових станцій: {str(e)}"
            messagebox.showerror("Помилка", error_msg)
            logging.error(error_msg)
            self.log_text.insert(tk.END, f"{error_msg}\n")
            self.log_text.see(tk.END)

    def _select_traffic_files(self):
        """Вибір файлів трафіку."""
        files = filedialog.askopenfilenames(
            title="Виберіть файли трафіку",
            filetypes=[
                ("Excel файли", "*.xlsx"),
                ("Всі файли", "*.*")
            ]
        )
        if files:
            self.traffic_files = files
            self.files_label.config(
                text=f"Вибрано файлів: {len(files)}"
            )
            self.log_text.insert(
                tk.END,
                f"Current Date and Time (UTC - YYYY-MM-DD HH:MM:SS formatted): "
                f"{datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"Current User's Login: {self.current_user}\n"
                f"Вибрано файли трафіку: {len(files)}\n\n"
            )
            self.log_text.see(tk.END)

    def _select_date_filter_file(self):
        """Вибір файлу з датами для фільтрації."""
        file = filedialog.askopenfilename(
            title="Виберіть файл з датами",
            filetypes=[
                ("Excel файли", "*.xlsx"),
                ("Всі файли", "*.*")
            ]
        )
        if file:
            self.date_filter_file = file
            self.log_text.insert(
                tk.END,
                f"Current Date and Time (UTC - YYYY-MM-DD HH:MM:SS formatted): "
                f"{datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"Current User's Login: {self.current_user}\n"
                f"Вибрано файл з датами: {file}\n\n"
            )
            self.log_text.see(tk.END)

    def _filter_by_date(self):
        """Фільтрація трафіку за датою."""
        try:
            if not hasattr(self, 'date_filter_file') or not self.date_filter_file:
                messagebox.showerror(
                    "Помилка",
                    "Спочатку виберіть файл з датами для фільтрації"
                )
                return

            if not self.traffic_files:
                messagebox.showerror(
                    "Помилка",
                    "Спочатку виберіть файли трафіку"
                )
                return

            # Отримуємо значення часового діапазону
            try:
                time_range = int(self.time_range.get())
                if time_range < 0:
                    raise ValueError("Значення повинно бути додатнім")
            except ValueError as e:
                messagebox.showerror(
                    "Помилка",
                    f"Некоректне значення часового діапазону: {str(e)}"
                )
                return

            # Створюємо директорію для результатів
            output_dir = os.path.join(
                os.path.dirname(self.date_filter_file),
                "results"
            )
            os.makedirs(output_dir, exist_ok=True)

            # Скидаємо прогрес-бар
            self.progress_bar['value'] = 0
            self.update_idletasks()

            # Викликаємо фільтрацію з однаковим діапазоном до і після події
            output_file, result_df = self.data_processor.filter_traffic_by_datetime(
                traffic_files=self.traffic_files,
                filter_file=self.date_filter_file,
                time_window_before=time_range,
                time_window_after=time_range,
                progress_bar=self.progress_bar,
                root=self,
                output_dir=output_dir
            )

            if output_file and result_df is not None:
                success_msg = (
                    f"Фільтрацію завершено успішно\n\n"
                    f"Знайдено записів: {len(result_df)}\n"
                    f"Збережено в файл:\n{os.path.basename(output_file)}"
                )
                messagebox.showinfo("Успіх", success_msg)

                # Додаємо запис в лог
                self.log_text.insert(
                    tk.END,
                    f"Фільтрацію за датою завершено. Знайдено записів: {len(result_df)}\n"
                )
                self.log_text.see(tk.END)
            else:
                messagebox.showwarning(
                    "Увага",
                    "Не знайдено даних за вказаними критеріями"
                )

        except Exception as e:
            error_msg = f"Помилка фільтрації за датою: {str(e)}"
            messagebox.showerror("Помилка", error_msg)
            logging.error(error_msg)

    def _select_geojson(self):
        """Вибір файлу GeoJSON з полігонами."""
        file = filedialog.askopenfilename(
            title="Виберіть файл GeoJSON",
            filetypes=[("GeoJSON файли", "*.geojson"), ("Всі файли", "*.*")]
        )
        if file:
            self.geojson_file = file
            self.log_text.insert(
                tk.END,
                f"Вибрано файл полігонів: {file}\n"
            )
            self.log_text.see(tk.END)

    @staticmethod
    def haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
        """
        Розрахунок відстані між двома точками за формулою гаверсинусів.

        Args:
            lat1: Широта першої точки
            lon1: Довгота першої точки
            lat2: Широта другої точки
            lon2: Довгота другої точки

        Returns:
            float: Відстань у метрах
        """
        R = 6371000  # радіус Землі в метрах
        lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
        c = 2 * atan2(sqrt(a), sqrt(1 - a))
        return R * c

    def extract_street_and_number(address):
        """
        Розділяє адресу на компоненти.

        Args:
            address (str): Повна адреса

        Returns:
            tuple: (регіон, район, населений пункт, вулиця, номер)
        """
        try:
            # Нормалізуємо адресу
            parts = address.split(',')
            parts = [p.strip() for p in parts]

            region = None
            district = None
            locality = None
            street = None
            number = None

            # Шукаємо компоненти
            for part in parts:
                part = part.upper()
                if 'ОБЛАСТЬ' in part or 'ОБЛ.' in part:
                    region = part
                elif 'РАЙОН' in part:
                    district = part
                elif any(x in part for x in ['М.', 'МІСТО', 'С.', 'СЕЛО', 'СМТ']):
                    locality = part
                elif any(x in part for x in ['ВУЛ.', 'ВУЛИЦЯ', 'БУЛ.', 'БУЛЬВАР', 'ПР.', 'ПРОСПЕКТ']):
                    # Виділяємо номер будинку, якщо є
                    match = re.search(r'(.+?)\s*(?:(?:БУД\.|БУДИНОК|Б\.)\s*)?(\d+(?:\/\d+)?)?$', part)
                    if match:
                        street = match.group(1).strip()
                        number = match.group(2)
                    else:
                        street = part

            return region, district, locality, street, number

        except Exception as e:
            logging.error(f"Помилка при розборі адреси {address}: {str(e)}")
            return None, None, None, None, None

    def find_closest_address(address, conn, threshold=90):
        """
        Знаходить найближчу адресу в базі даних.

        Args:
            address (str): Адреса для пошуку
            conn: З'єднання з базою даних
            threshold (int): Поріг схожості (за замовчуванням 90)

        Returns:
            tuple: (широта, довгота) або None якщо не знайдено
        """
        try:
            normalized_address = ' '.join(str(address).split()).upper()
            traffic_region, traffic_district, traffic_locality, traffic_street, traffic_number = extract_street_and_number(
                normalized_address)

            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM addresses")
            if cursor.fetchone()[0] == 0:
                logging.error("База адресов пуста")
                return None

            query = "SELECT address, latitude, longitude FROM addresses WHERE address LIKE ?"
            search_pattern = f"%{traffic_street or normalized_address}%"
            cursor.execute(query, (search_pattern,))
            candidates = cursor.fetchall()

            if not candidates:
                logging.warning(f"Для адреса {address} не найдено кандидатов")
                return None

            best_match = None
            best_score = 0
            best_coords = None

            for candidate in candidates:
                db_address, latitude, longitude = candidate
                db_region, db_district, db_locality, db_street, db_number = extract_street_and_number(db_address)

                score = fuzz.ratio(normalized_address, db_address)
                if score >= threshold and score > best_score:
                    # Перевірка компонентів адреси
                    if traffic_street and db_street and traffic_street != db_street:
                        continue
                    if traffic_number and db_number and traffic_number != db_number:
                        continue
                    if traffic_number is None and db_number is not None:
                        continue
                    if db_number is None and traffic_number is not None:
                        continue
                    if traffic_locality and db_locality and traffic_locality != db_locality:
                        continue
                    if traffic_region and db_region and traffic_region != db_region:
                        continue
                    if traffic_district and db_district and traffic_district != db_district:
                        continue

                    best_score = score
                    best_match = db_address
                    best_coords = (latitude, longitude)

            if best_match:
                logging.info(f"Лучший матч для {address}: {best_match} (score={best_score})")

            return best_coords

        except Exception as e:
            logging.error(f"Помилка пошуку адреси {address}: {str(e)}")
            return None

    def load_address_coords_from_db(self):
        """
        Завантажує словник координат з SQLite бази даних.
        """
        unique_address_coords = {}
        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'addresses.db')

        try:
            if not os.path.exists(db_path):
                logging.error(f"База даних не знайдена: {db_path}")
                return unique_address_coords

            conn = sqlite3.connect(db_path)

            # Для кожної адреси спробуємо знайти найближчу відповідність
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT address FROM addresses")
            addresses = cursor.fetchall()

            for (address,) in addresses:
                coords = find_closest_address(address, conn)
                if coords:
                    unique_address_coords[str(address).strip()] = coords

            logging.info(f"Завантажено {len(unique_address_coords)} адрес з бази даних")

            conn.close()
            return unique_address_coords

        except Exception as e:
            logging.error(f"Помилка при завантаженні координат з бази даних: {str(e)}")
            return unique_address_coords

    def create_temp_traffic_db(self, meetings_files):
        """
        Створення тимчасової бази для зустрічей.

        Args:
            meetings_files: Список файлів для обробки

        Returns:
            sqlite3.Connection: З'єднання з базою даних
        """
        conn = sqlite3.connect(':memory:')
        cursor = conn.cursor()

        # Створюємо структуру бази
        cursor.execute('''
            CREATE TABLE traffic (
                subscriber_a TEXT,
                date TEXT,
                time TEXT,
                azimuth REAL,
                address TEXT,
                latitude REAL,
                longitude REAL
            )
        ''')

        # Створюємо індекси для оптимізації
        cursor.execute('CREATE INDEX idx_subscriber ON traffic (subscriber_a)')
        cursor.execute('CREATE INDEX idx_date_time ON traffic (date, time)')
        cursor.execute('CREATE INDEX idx_coords ON traffic (latitude, longitude)')

        # Додаємо дані з файлів
        for file in meetings_files:
            if not os.path.exists(file):
                logging.warning(f"Файл {file} не знайдено")
                continue

            df = pd.read_excel(file)
            required_columns = ['Абонент А', 'Дата', 'Час', 'Адреса БС', 'Широта', 'Долгота']

            if not all(col in df.columns for col in required_columns):
                logging.error(f"У файлі {file} відсутні необхідні стовпці")
                continue

            has_azimuth = 'Аз.' in df.columns

            for _, row in df.iterrows():
                azimuth = row['Аз.'] if has_azimuth and pd.notna(row['Аз.']) else None
                cursor.execute('''
                    INSERT INTO traffic (subscriber_a, date, time, azimuth, address, latitude, longitude)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    row['Абонент А'],
                    str(row['Дата']),
                    row['Час'],
                    azimuth,
                    row['Адреса БС'],
                    row['Широта'],
                    row['Долгота']
                ))

        conn.commit()
        logging.info(f"Створено тимчасову базу з {len(meetings_files)} файлів")
        return conn

    def find_meetings_sql(self, conn, max_distance=400, time_delta_minutes=30, output_dir=None):
        """
        Пошук зустрічей в базі даних.

        Args:
            conn: З'єднання з базою даних
            max_distance: Максимальна відстань між точками в метрах
            time_delta_minutes: Часове вікно в хвилинах
            output_dir: Тека для збереження результатів

        Returns:
            tuple: (список зустрічей, шлях до файлу результатів)
        """
        cursor = conn.cursor()
        meetings = []
        time_delta_seconds = time_delta_minutes * 60

        query = '''
            SELECT 
                t1.subscriber_a AS subscriber_a1,
                t2.subscriber_a AS subscriber_a2,
                t1.date AS date,
                t1.time AS time1,
                t1.azimuth AS azimuth1,
                t1.address AS address1,
                t1.latitude AS lat1,
                t1.longitude AS lon1,
                t2.time AS time2,
                t2.azimuth AS azimuth2,
                t2.address AS address2,
                t2.latitude AS lat2,
                t2.longitude AS lon2
            FROM traffic t1
            JOIN traffic t2
            ON t1.subscriber_a < t2.subscriber_a
            AND t1.date = t2.date
            WHERE ABS(
                (CAST(strftime('%s', t1.time) AS INTEGER) - CAST(strftime('%s', t2.time) AS INTEGER))
            ) <= ?
            AND ? * (
                6371000 * acos(
                    cos(radians(t1.latitude)) * cos(radians(t2.latitude)) * 
                    cos(radians(t2.longitude) - radians(t1.longitude)) + 
                    sin(radians(t1.latitude)) * sin(radians(t2.latitude))
                )
            ) <= ?
        '''

        cursor.execute(query, (time_delta_seconds, 1, max_distance))

        for row in cursor.fetchall():
            subscriber_a1, subscriber_a2, date, time1, azimuth1, address1, lat1, lon1, \
                time2, azimuth2, address2, lat2, lon2 = row

            distance = self.haversine_distance(lat1, lon1, lat2, lon2)

            if distance <= max_distance:
                meetings.append({
                    'Абонент А 1': subscriber_a1,
                    'Абонент А 2': subscriber_a2,
                    'Дата': date,
                    'Время 1': time1,
                    'Азимут 1': azimuth1,
                    'Адрес 1': address1,
                    'Широта 1': lat1,
                    'Долгота 1': lon1,
                    'Время 2': time2,
                    'Азимут 2': azimuth2,
                    'Адрес 2': address2,
                    'Широта 2': lat2,
                    'Долгота 2': lon2,
                    'Расстояние (м)': round(distance, 2)
                })

        # Збереження результатів у файл
        meetings_file = None
        if meetings and output_dir:
            try:
                # Перевірка, чи існує директорія
                if not os.path.exists(output_dir):
                    os.makedirs(output_dir)
                    logging.info(f"Створено директорію: {output_dir}")

                # Створення імені файлу з часовою міткою
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                meetings_file = os.path.join(output_dir, f"meetings_{timestamp}.xlsx")

                # Перетворення списку зустрічей у DataFrame
                meetings_df = pd.DataFrame(meetings)

                # Збереження у файл Excel
                meetings_df.to_excel(meetings_file, index=False)
                logging.info(f"Результати зустрічей збережено у файл: {meetings_file}")

            except Exception as e:
                logging.error(f"Помилка збереження файлу зустрічей: {str(e)}")
                meetings_file = None

        return meetings, meetings_file

    def find_meetings(self):
        """Пошук зустрічей."""
        try:
            if not self.traffic_files:
                raise ValueError("Не вибрано файли трафіку")

            # Визначаємо теку першого файлу для збереження результатів
            output_dir = os.path.dirname(self.traffic_files[0])

            # Отримуємо параметри
            try:
                max_distance = float(self.max_distance.get())
                time_window = int(self.time_window.get())
            except ValueError:
                raise ValueError(
                    "Неправильний формат відстані або часового вікна"
                )

            # Створюємо тимчасову базу даних і шукаємо зустрічі
            conn = self.create_temp_traffic_db(self.traffic_files)
            meetings, meetings_file = self.find_meetings_sql(
                conn,
                max_distance=max_distance,
                time_delta_minutes=time_window,
                output_dir=output_dir
            )

            if meetings and meetings_file:
                self.log_text.insert(
                    tk.END,
                    f"{self._get_current_datetime_and_user()}\n"
                    f"Знайдено {len(meetings)} зустрічей\n"
                    f"Результат збережено: {meetings_file}\n"
                )
            else:
                self.log_text.insert(
                    tk.END,
                    f"{self._get_current_datetime_and_user()}\n"
                    "Зустрічей не знайдено\n"
                )
            self.log_text.see(tk.END)

        except Exception as e:
            messagebox.showerror("Помилка", str(e))
            logging.error(f"Помилка пошуку зустрічей: {e}")
            self.log_text.insert(
                tk.END,
                f"{self._get_current_datetime_and_user()}\n"
                f"Помилка: {str(e)}\n"
            )
            self.log_text.see(tk.END)

    def _process_files(self) -> None:
        """Обробка файлів для аналізу переміщень."""
        try:
            if not self.traffic_files:
                raise ValueError("Не вибрано файли трафіку")

            # Отримуємо параметри
            try:
                max_distance = float(self.max_distance.get())
                day_min_duration = int(self.day_min_duration.get())
                night_min_duration = int(self.night_min_duration.get())
            except ValueError as e:
                raise ValueError("Неправильний формат параметрів") from e

            # Створюємо директорію для результатів
            output_dir = os.path.join(
                os.path.dirname(self.traffic_files[0]),
                "results"
            )
            os.makedirs(output_dir, exist_ok=True)

            # Ініціалізуємо список для адрес без координат
            no_coords_data = []

            # Завантажуємо словник координат з бази даних
            unique_address_coords = self.load_address_coords_from_db()

            if not unique_address_coords:
                logging.warning("Не знайдено координат в базі даних")

            # Обробляємо кожен файл
            processed_files = []
            for file in self.traffic_files:
                try:
                    # Викликаємо process_traffic_file як глобальну функцію
                    output_file, polygons_file = process_traffic_file(
                        file,
                        unique_address_coords,
                        self.geojson_file if hasattr(self, 'geojson_file') else None,
                        no_coords_data,
                        self.progress_bar,
                        self,
                        output_dir
                    )

                    if output_file:
                        processed_files.append(output_file)

                except Exception as e:
                    logging.error(f"Помилка обробки файлу {file}: {str(e)}")
                    continue

            # Зберігаємо список адрес без координат
            if self.no_coords_data:
                output_dir = os.path.dirname(self.traffic_files[0])
                self._save_no_coords_file(output_dir)
                self.log_text.insert(
                    tk.END,
                    f"Збережено список адрес без координат\n"
                )

            # Виводимо підсумок
            self.log_text.insert(
                tk.END,
                f"\n{self._get_current_datetime_and_user()}\n"
                f"Обробка завершена\n"
                f"Успішно оброблено файлів: {len(processed_files)}/{len(self.traffic_files)}\n"
                f"Адрес без координат: {len(self.no_coords_data)}\n"
            )
            self.log_text.see(tk.END)

        except Exception as e:
            messagebox.showerror("Помилка", str(e))
            logging.error(f"Помилка обробки файлів: {e}")

    def _get_current_datetime_and_user(self) -> str:
        """
        Отримати поточні дату, час та користувача.

        Returns:
            str: Рядок з датою, часом та користувачем
        """
        try:
            return f"Час: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}, користувач: {os.getlogin()}"
        except Exception as e:
            logging.error(f"Помилка отримання дати та користувача: {e}")
            return ""

    def _save_no_coords_file(self, output_dir: str):
        """
        Збереження списку адрес без координат.

        Args:
            output_dir: Тека для збереження файлу
        """
        try:
            no_coords_file = os.path.join(output_dir, "БС без координат.xlsx")
            if self.no_coords_data:
                no_coords_df = pd.DataFrame(self.no_coords_data)
                no_coords_df = no_coords_df.drop_duplicates(subset=['Адреса БС'])
                no_coords_df.to_excel(no_coords_file, index=False)
                logging.info(
                    f"Збережено файл без координат (без дублікатів): "
                    f"{no_coords_file}, записів: {len(no_coords_df)}"
                )
            else:
                no_coords_df = pd.DataFrame([{'Адреса БС': 'Всі адреси мають координати'}])
                no_coords_df.to_excel(no_coords_file, index=False)
                logging.info(f"Збережено файл без координат: {no_coords_file}")
        except Exception as e:
            logging.error(f"Помилка збереження файлу без координат: {e}")
            raise

    def _merge_files(self):
        """Об'єднання файлів трафіку."""
        try:
            if not self.traffic_files:
                raise ValueError("Не вибрано файли трафіку")

            # Визначаємо теку першого файлу для збереження результатів
            self.output_dir = os.path.dirname(self.traffic_files[0])

            # Отримуємо дозволені типи з'єднань
            allowed_types = []
            if self.connection_type.get() != "Всі":
                allowed_types = [self.connection_type.get()]
            else:
                allowed_types = ["вих", "вих СМС", "вх", "вх СМС", "переадр"]

            # Об'єднуємо файли
            output_files, result_df = self.data_processor.merge_traffic_files(
                self.traffic_files,
                "Тип",  # назва колонки типу з'єднання
                allowed_types,
                self.progress_bar,
                self.winfo_toplevel(),
                self.output_dir
            )

            if output_files:
                message = f"Файли успішно об'єднано:\n"
                for file in output_files:
                    message += f"- {file}\n"
                self.log_text.insert(tk.END, message)
            else:
                self.log_text.insert(
                    tk.END,
                    "Помилка об'єднання файлів\n"
                )
            self.log_text.see(tk.END)

        except Exception as e:
            messagebox.showerror("Помилка", str(e))
            logging.error(f"Помилка об'єднання файлів: {e}")

    def _analyze_activity(self):
        """Аналіз активності по днях та місяцях."""
        try:
            if not self.traffic_files:
                raise ValueError("Не вибрано файли трафіку")

            # Створюємо вікно для графіка
            graph_window = tk.Toplevel(self)
            graph_window.title("Аналіз активності")
            graph_window.geometry("1200x800")

            # Створюємо фрейм з прокруткою
            main_frame = ttk.Frame(graph_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

            # Додаємо скролбар
            canvas = tk.Canvas(main_frame)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )

            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            # Створюємо DataFrame для аналізу
            all_data = []
            for file in self.traffic_files:
                df = pd.read_excel(file)
                df['Дата'] = pd.to_datetime(df['Дата'], format='%d.%m.%Y', dayfirst=True)
                all_data.append(df)

            combined_df = pd.concat(all_data)

            # Додаємо місяць та рік до даних
            combined_df['Місяць_Рік'] = combined_df['Дата'].dt.strftime('%Y-%m')
            combined_df['Година'] = pd.to_datetime(combined_df['Час'], format='%H:%M:%S').dt.hour

            # Групуємо дані по місяцях
            months = sorted(combined_df['Місяць_Рік'].unique())

            # Для кожного місяця створюємо окремий графік
            for month in months:
                month_data = combined_df[combined_df['Місяць_Рік'] == month]

                # Створюємо фрейм для місяця
                month_frame = ttk.LabelFrame(scrollable_frame, text=f"Активність за {month}")
                month_frame.pack(fill=tk.X, padx=5, pady=5)

                # Створюємо фігуру з двома графіками
                fig = Figure(figsize=(12, 6))

                # Графік активності по днях
                daily_activity = month_data.groupby('Дата').size()
                ax1 = fig.add_subplot(121)
                daily_activity.plot(kind='bar', ax=ax1)
                ax1.set_title('Активність по днях')
                ax1.set_xlabel('Дата')
                ax1.set_ylabel('Кількість подій')
                plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45)

                # Графік активності по годинах
                hourly_activity = month_data.groupby('Година').size()
                ax2 = fig.add_subplot(122)
                hourly_activity.plot(kind='bar', ax=ax2)
                ax2.set_title('Активність по годинах')
                ax2.set_xlabel('Година')
                ax2.set_ylabel('Кількість подій')

                # Додаємо графіки у вікно
                canvas_widget = FigureCanvasTkAgg(fig, master=month_frame)
                canvas_widget.draw()
                canvas_widget.get_tk_widget().pack(fill=tk.BOTH, expand=True)

                # Додаємо статистику для місяця
                stats_frame = ttk.Frame(month_frame)
                stats_frame.pack(fill=tk.X, padx=5, pady=5)

                total_events = len(month_data)
                unique_dates = len(daily_activity)
                avg_daily = total_events / unique_dates if unique_dates > 0 else 0
                peak_hour = hourly_activity.idxmax() if not hourly_activity.empty else 0
                peak_day = daily_activity.idxmax() if not daily_activity.empty else None

                stats_text = (
                    f"Загальна кількість подій: {total_events}\n"
                    f"Кількість днів: {unique_dates}\n"
                    f"Середня кількість подій за день: {avg_daily:.2f}\n"
                    f"Пікова година: {peak_hour}:00\n"
                    f"День з найбільшою активністю: {peak_day.strftime('%d.%m.%Y') if peak_day else 'Немає даних'}"
                )

                ttk.Label(stats_frame, text=stats_text).pack(padx=5, pady=5)

            # Налаштування прокрутки
            main_frame.pack(fill=tk.BOTH, expand=True)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # Додаємо інформацію в лог
            self.log_text.insert(
                tk.END,
                f"{self._get_current_datetime_and_user()}\n"
                f"Проведено аналіз активності\n"
                f"Оброблено файлів: {len(self.traffic_files)}\n"
                f"Проаналізовано місяців: {len(months)}\n"
            )
            self.log_text.see(tk.END)

        except Exception as e:
            messagebox.showerror("Помилка", str(e))
            logging.error(f"Помилка аналізу активності: {e}")
            self.log_text.insert(
                tk.END,
                f"{self._get_current_datetime_and_user()}\n"
                f"Помилка: {str(e)}\n"
            )
            self.log_text.see(tk.END)

    def _analyze_specific_date(self):
        """Аналіз активності за конкретний день."""
        try:
            if not self.traffic_files:
                raise ValueError("Не вибрано файли трафіку")

            date_str = self.specific_date.get().strip()
            if not date_str:
                raise ValueError("Введіть дату")

            try:
                specific_date = datetime.strptime(date_str, '%d.%m.%Y')
            except ValueError:
                raise ValueError("Неправильний формат дати. Використовуйте DD.MM.YYYY")

            # Створюємо вікно для графіка
            graph_window = tk.Toplevel(self)
            graph_window.title(f"Аналіз активності за {date_str}")
            graph_window.geometry("800x600")

            # Створюємо DataFrame для аналізу
            all_data = []
            for file in self.traffic_files:
                df = pd.read_excel(file)
                df['Дата'] = pd.to_datetime(df['Дата'], format='%d.%m.%Y', dayfirst=True)
                df['Година'] = pd.to_datetime(df['Час'], format='%H:%M:%S').dt.hour
                df['Хвилина'] = pd.to_datetime(df['Час'], format='%H:%M:%S').dt.minute
                all_data.append(df)

            combined_df = pd.concat(all_data)

            # Фільтруємо дані за вказану дату
            day_data = combined_df[combined_df['Дата'].dt.date == specific_date.date()]

            if len(day_data) == 0:
                raise ValueError(f"Немає даних за {date_str}")

            # Створюємо фігуру з графіками
            fig = Figure(figsize=(12, 8))

            # Графік активності по годинах
            hourly_activity = day_data.groupby('Година').size()
            ax1 = fig.add_subplot(211)
            hourly_activity.plot(kind='bar', ax=ax1)
            ax1.set_title(f'Активність по годинах за {date_str}')
            ax1.set_xlabel('Година')
            ax1.set_ylabel('Кількість подій')

            # Детальний графік по годинах та хвилинах
            ax2 = fig.add_subplot(212)
            day_data['Час_години'] = day_data['Година'] + day_data['Хвилина'] / 60
            time_series = pd.Series(index=day_data['Час_години'], data=1)
            time_series = time_series.sort_index()
            ax2.plot(time_series.index, np.ones_like(time_series.index), '|', markersize=10)
            ax2.set_title('Розподіл подій протягом дня')
            ax2.set_xlabel('Година')
            ax2.set_xlim(-0.5, 23.5)
            ax2.grid(True)

            # Додаємо графіки у вікно
            canvas = FigureCanvasTkAgg(fig, master=graph_window)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

            # Додаємо статистику
            stats_frame = ttk.LabelFrame(graph_window, text="Статистика")
            stats_frame.pack(fill=tk.X, padx=5, pady=5)

            # Розраховуємо статистику
            total_events = len(day_data)
            unique_hours = len(hourly_activity)
            peak_hour = hourly_activity.idxmax()
            peak_hour_count = hourly_activity.max()
            avg_per_hour = total_events / 24

            # Знаходимо найдовший проміжок без подій
            times = sorted(day_data['Час_години'])
            max_gap = 0
            gap_start = 0
            if times:
                gaps = [(times[i + 1] - times[i], times[i]) for i in range(len(times) - 1)]
                if gaps:
                    max_gap, gap_start = max(gaps)

            stats_text = (
                f"Загальна кількість подій: {total_events}\n"
                f"Активних годин: {unique_hours}\n"
                f"Середня кількість подій за годину: {avg_per_hour:.2f}\n"
                f"Пікова година: {peak_hour}:00 ({peak_hour_count} подій)\n"
                f"Найдовший проміжок без подій: {max_gap:.2f} годин (початок: {int(gap_start)}:00)"
            )

            ttk.Label(stats_frame, text=stats_text).pack(padx=5, pady=5)

            # Додаємо інформацію в лог
            self.log_text.insert(
                tk.END,
                f"{self._get_current_datetime_and_user()}\n"
                f"Проведено аналіз активності за {date_str}\n"
                f"{stats_text}\n"
            )
            self.log_text.see(tk.END)

        except Exception as e:
            messagebox.showerror("Помилка", str(e))
            logging.error(f"Помилка аналізу активності за день: {e}")
            self.log_text.insert(
                tk.END,
                f"{self._get_current_datetime_and_user()}\n"
                f"Помилка: {str(e)}\n"
            )
            self.log_text.see(tk.END)

    def check_sectors_overlap(self, lat1, lon1, azimuth1, lat2, lon2, azimuth2, radius, angle):
        """
        Перевірка перетину двох секторів.

        Args:
            lat1, lon1: координати першої точки
            azimuth1: азимут першого сектора
            lat2, lon2: координати другої точки
            azimuth2: азимут другого сектора
            radius: радіус секторів в метрах
            angle: кут секторів в градусах

        Returns:
            bool: True якщо сектори перетинаються
        """
        # Розрахунок відстані між точками
        distance = self.haversine_distance(lat1, lon1, lat2, lon2)

        # Якщо відстань більша за подвійний радіус, сектори точно не перетинаються
        if distance > radius * 2:
            return False

        # Розрахунок кутів секторів
        half_angle = angle / 2

        # Розрахунок напрямку від точки 1 до точки 2
        y = sin(radians(lon2 - lon1)) * cos(radians(lat2))
        x = cos(radians(lat1)) * sin(radians(lat2)) - \
            sin(radians(lat1)) * cos(radians(lat2)) * cos(radians(lon2 - lon1))
        bearing = (degrees(atan2(y, x)) + 360) % 360

        # Перевірка чи потрапляє точка 2 в сектор точки 1
        angle_diff1 = abs((azimuth1 - bearing + 180) % 360 - 180)
        in_sector1 = angle_diff1 <= half_angle

        # Перевірка чи потрапляє точка 1 в сектор точки 2
        bearing_back = (bearing + 180) % 360
        angle_diff2 = abs((azimuth2 - bearing_back + 180) % 360 - 180)
        in_sector2 = angle_diff2 <= half_angle

        return in_sector1 and in_sector2

    def find_meetings_sql(self, conn, max_distance=400, time_delta_minutes=30, output_dir=None):
        """
        Пошук зустрічей в базі даних з урахуванням секторів.
        """
        try:
            cursor = conn.cursor()
            meetings = []
            time_delta_seconds = time_delta_minutes * 60

            # Отримуємо параметри секторів
            sector_radius = float(self.sector_radius.get())
            sector_angle = float(self.sector_angle.get())

            query = '''
                SELECT 
                    t1.subscriber_a AS subscriber_a1,
                    t2.subscriber_a AS subscriber_a2,
                    t1.date AS date,
                    t1.time AS time1,
                    t1.azimuth AS azimuth1,
                    t1.address AS address1,
                    t1.latitude AS lat1,
                    t1.longitude AS lon1,
                    t2.time AS time2,
                    t2.azimuth AS azimuth2,
                    t2.address AS address2,
                    t2.latitude AS lat2,
                    t2.longitude AS lon2
                FROM traffic t1
                JOIN traffic t2
                ON t1.subscriber_a < t2.subscriber_a
                AND t1.date = t2.date
                WHERE ABS(
                    (CAST(strftime('%s', t1.time) AS INTEGER) - CAST(strftime('%s', t2.time) AS INTEGER))
                ) <= ?
                AND (
                    ? * (
                        6371000 * acos(
                            cos(radians(t1.latitude)) * cos(radians(t2.latitude)) * 
                            cos(radians(t2.longitude) - radians(t1.longitude)) + 
                            sin(radians(t1.latitude)) * sin(radians(t2.latitude))
                        )
                    ) <= ?
                    OR (
                        t1.azimuth IS NOT NULL 
                        AND t2.azimuth IS NOT NULL
                        AND ? * (
                            6371000 * acos(
                                cos(radians(t1.latitude)) * cos(radians(t2.latitude)) * 
                                cos(radians(t2.longitude) - radians(t1.longitude)) + 
                                sin(radians(t1.latitude)) * sin(radians(t2.latitude))
                            )
                        ) <= ?
                    )
                )
            '''

            # Виконуємо запит з більшою максимальною відстанню для перевірки секторів
            cursor.execute(query, (
                time_delta_seconds,
                1, max_distance,
                1, sector_radius * 2
            ))

            for row in cursor.fetchall():
                subscriber_a1, subscriber_a2, date, time1, azimuth1, address1, lat1, lon1, \
                    time2, azimuth2, address2, lat2, lon2 = row

                distance = self.haversine_distance(lat1, lon1, lat2, lon2)

                # Перевіряємо умови зустрічі
                is_meeting = False

                # Звичайна перевірка відстані
                if distance <= max_distance:
                    is_meeting = True
                    meeting_type = "Відстань"
                # Перевірка перетину секторів
                elif (azimuth1 is not None and azimuth2 is not None and
                      self.check_sectors_overlap(
                          lat1, lon1, azimuth1,
                          lat2, lon2, azimuth2,
                          sector_radius, sector_angle
                      )):
                    is_meeting = True
                    meeting_type = "Сектори"

                if is_meeting:
                    meetings.append({
                        'Абонент А 1': subscriber_a1,
                        'Абонент А 2': subscriber_a2,
                        'Дата': date,
                        'Час 1': time1,
                        'Азимут 1': azimuth1,
                        'Адреса 1': address1,
                        'Широта 1': lat1,
                        'Довгота 1': lon1,
                        'Час 2': time2,
                        'Азимут 2': azimuth2,
                        'Адреса 2': address2,
                        'Широта 2': lat2,
                        'Довгота 2': lon2,
                        'Відстань (м)': round(distance, 2),
                        'Тип збігу': meeting_type
                    })

            # Збереження результатів у файл
            meetings_file = None
            if meetings and output_dir:
                try:
                    if not os.path.exists(output_dir):
                        os.makedirs(output_dir)

                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    meetings_file = os.path.join(output_dir, f"meetings_{timestamp}.xlsx")

                    meetings_df = pd.DataFrame(meetings)
                    meetings_df.to_excel(meetings_file, index=False)

                    # Створюємо карту зустрічей
                    map_file = os.path.join(output_dir, f"meetings_map_{timestamp}.html")
                    self.create_meetings_map_with_excel(meetings, map_file)

                    logging.info(f"Результати зустрічей збережено у файл: {meetings_file}")

                except Exception as e:
                    logging.error(f"Помилка збереження файлу зустрічей: {str(e)}")
                    meetings_file = None

            return meetings, meetings_file

        except Exception as e:
            logging.error(f"Помилка пошуку зустрічей: {str(e)}")
            raise

    def create_meetings_map_with_excel(self, meetings, filename):
        """
        Створення окремих карт зустрічей для секторів та відстаней з контролем відображення.
        Карти будуть показувати всі дані без фільтрації по датах.
        Дублікати секторів в одну дату будуть відображені лише один раз.

        ДОДАНИЙ ВИКЛИК create_excel_traffic_reports для створення файлів трафіку!
        """
        import folium
        import numpy as np
        from datetime import datetime
        import logging
        import tkinter as tk
        import math

        def _get_current_datetime_and_user():
            return "2025-07-21 10:28:49 - McNeal1994"

        def parse_datetime(date_str):
            """Парсить дату з формату DD.MM.YYYY та повертає 'YYYY-MM-DD'."""
            try:
                return datetime.strptime(date_str, '%d.%m.%Y').strftime('%Y-%m-%d')
            except ValueError:
                logging.error(f"Помилка парсингу дати: {date_str}")
                return None

        def get_base_station_name(meeting_data, idx):
            """Отримує назву базової станції з поля 'Адреса 1' або 'Адреса 2'"""
            # Правильні назви полів
            address_field = f'Адреса {idx}'

            if address_field in meeting_data and meeting_data[address_field]:
                return meeting_data[address_field].strip()

            # Якщо не знайдено, генеруємо на основі координат
            try:
                lat = float(meeting_data.get(f'Широта {idx}', 0))
                lon = float(meeting_data.get(f'Довгота {idx}', 0))
                lat_str = f"{abs(lat):.4f}{'N' if lat >= 0 else 'S'}"
                lon_str = f"{abs(lon):.4f}{'E' if lon >= 0 else 'W'}"
                return f"БС-{lat_str}-{lon_str}"
            except (ValueError, TypeError):
                return "БС-Невідома"

        def create_sector_polygon(lat, lon, azimuth, radius, angle):
            """Створює полігон сектора для відображення на карті"""
            # Конвертуємо в радіани
            azimuth_rad = math.radians(azimuth)
            half_angle_rad = math.radians(angle / 2)

            # Точки сектора
            points = [[lat, lon]]  # Центральна точка

            # Створюємо точки по дузі сектора
            num_points = 20  # Кількість точок для плавної дуги
            for i in range(num_points + 1):
                angle_step = (2 * half_angle_rad * i) / num_points - half_angle_rad
                current_azimuth = azimuth_rad + angle_step

                # Обчислюємо координати точки на дузі
                lat_offset = (radius / 111000) * math.cos(current_azimuth)  # ~111км на градус широти
                lon_offset = (radius / (111000 * math.cos(math.radians(lat)))) * math.sin(current_azimuth)

                new_lat = lat + lat_offset
                new_lon = lon + lon_offset
                points.append([new_lat, new_lon])

            points.append([lat, lon])  # Закриваємо полігон
            return points

        def create_date_groups(map_obj, dates):
            """Створює групи для дат на карті"""
            return {
                d: folium.FeatureGroup(
                    name=datetime.strptime(d, '%Y-%m-%d').strftime('%d.%m.%Y'),
                    show=True
                ).add_to(map_obj)
                for d in dates
            }

        def add_sectors_to_map(map_obj, sectors, groups):
            """Додає сектори на карту з уникненням дублікатів"""
            # Словники для відстеження унікальних елементів по датах
            unique_sectors_by_date = {}
            unique_markers_by_date = {}

            # Спочатку додаємо всі сектори (щоб вони були позаду)
            for m in sectors:
                d = parse_datetime(m.get('Дата', ''))
                if not d or d not in groups:
                    continue

                # Ініціалізуємо структури для дати
                if d not in unique_sectors_by_date:
                    unique_sectors_by_date[d] = set()

                grp = groups[d]

                # Обробляємо кожен сектор окремо
                for idx, color in [('1', 'blue'), ('2', 'red')]:
                    try:
                        lat = float(m.get(f'Широта {idx}', 0))
                        lon = float(m.get(f'Довгота {idx}', 0))
                        azimuth = float(m.get(f'Азимут {idx}', 0))

                        # Створюємо унікальний ключ для сектора (округлюємо координати до 6 знаків)
                        sector_key = (round(lat, 6), round(lon, 6), round(azimuth, 1))

                        # Перевіряємо, чи не додавали вже цей сектор в цю дату
                        if sector_key not in unique_sectors_by_date[d]:
                            unique_sectors_by_date[d].add(sector_key)

                            # Отримуємо радіус та кут сектора
                            radius = float(self.sector_radius.get()) if hasattr(self, 'sector_radius') else 1000
                            sector_angle = float(self.sector_angle.get()) if hasattr(self, 'sector_angle') else 60

                            # Створюємо полігон сектора
                            sector_points = create_sector_polygon(lat, lon, azimuth, radius, sector_angle)

                            # Отримуємо назву БС для сектора
                            bs_name = get_base_station_name(m, idx)
                            abonent = m.get(f'Абонент А {idx}', 'Невідомо')

                            # Додаємо сектор як полігон
                            folium.Polygon(
                                locations=sector_points,
                                color=color,
                                weight=2,
                                opacity=0.7,
                                fill=True,
                                fillColor=color,
                                fillOpacity=0.3,
                                popup=folium.Popup(
                                    f"<b>Абонент:</b> {abonent}<br>"
                                    f"<b>Адреса БС:</b> {bs_name}<br>"
                                    f"<b>Азимут:</b> {azimuth}°",
                                    max_width=250,
                                    sticky=True
                                )
                            ).add_to(grp)

                    except (ValueError, TypeError) as e:
                        logging.error(f"Помилка обробки координат для сектора {idx}: {e}")
                        continue

            # Потім додаємо лінії між точками
            for m in sectors:
                d = parse_datetime(m.get('Дата', ''))
                if not d or d not in groups:
                    continue

                grp = groups[d]

                try:
                    lat1 = float(m.get('Широта 1', 0))
                    lon1 = float(m.get('Довгота 1', 0))
                    lat2 = float(m.get('Широта 2', 0))
                    lon2 = float(m.get('Довгота 2', 0))

                    # Унікальний ключ для лінії
                    line_key = (round(min(lat1, lat2), 6), round(min(lon1, lon2), 6),
                                round(max(lat1, lat2), 6), round(max(lon1, lon2), 6))

                    # Ініціалізуємо словник для ліній, якщо потрібно
                    if not hasattr(add_sectors_to_map, 'unique_lines_by_date'):
                        add_sectors_to_map.unique_lines_by_date = {}
                    if d not in add_sectors_to_map.unique_lines_by_date:
                        add_sectors_to_map.unique_lines_by_date[d] = set()

                    # Додаємо лінію тільки якщо її ще немає
                    if line_key not in add_sectors_to_map.unique_lines_by_date[d]:
                        add_sectors_to_map.unique_lines_by_date[d].add(line_key)

                        folium.PolyLine(
                            locations=[[lat1, lon1], [lat2, lon2]],
                            color='purple', weight=3, opacity=0.8,
                            popup=folium.Popup(
                                f"<b>Зв'язок між секторами</b>",
                                max_width=200,
                                sticky=True
                            )
                        ).add_to(grp)

                except (ValueError, TypeError) as e:
                    logging.error(f"Помилка створення лінії між секторами: {e}")

            # В кінці додаємо маркери (щоб були найвище, над секторами)
            for m in sectors:
                d = parse_datetime(m.get('Дата', ''))
                if not d or d not in groups:
                    continue

                # Ініціалізуємо структури для дати
                if d not in unique_markers_by_date:
                    unique_markers_by_date[d] = set()

                grp = groups[d]

                for idx, color in [('1', 'blue'), ('2', 'red')]:
                    try:
                        lat = float(m.get(f'Широта {idx}', 0))
                        lon = float(m.get(f'Довгота {idx}', 0))
                        azimuth = float(m.get(f'Азимут {idx}', 0))

                        # Створюємо унікальний ключ для маркера
                        marker_key = (round(lat, 6), round(lon, 6))

                        # Перевіряємо, чи не додавали вже цей маркер в цю дату
                        if marker_key not in unique_markers_by_date[d]:
                            unique_markers_by_date[d].add(marker_key)

                            # Збираємо інформацію про абонентів на цій точці
                            same_location_info = []

                            for m2 in sectors:
                                d2 = parse_datetime(m2.get('Дата', ''))
                                if d2 == d:  # Тільки для тієї ж дати
                                    for idx2 in ['1', '2']:
                                        lat2 = float(m2.get(f'Широта {idx2}', 0))
                                        lon2 = float(m2.get(f'Довгота {idx2}', 0))
                                        if abs(lat2 - lat) < 0.000001 and abs(lon2 - lon) < 0.000001:
                                            azimuth2 = float(m2.get(f'Азимут {idx2}', 0))
                                            abonent = m2.get(f'Абонент А {idx2}', 'Невідомо')
                                            bs_name = get_base_station_name(m2, idx2)

                                            # Уникаємо дублікатів
                                            info_key = f"{abonent}_{azimuth2}_{bs_name}"
                                            if info_key not in [f"{info['абонент']}_{info['азимут']}_{info['адреса']}"
                                                                for info in same_location_info]:
                                                same_location_info.append({
                                                    'абонент': abonent,
                                                    'адреса': bs_name,
                                                    'азимут': azimuth2
                                                })

                            # Формуємо інформацію для popup
                            popup_info = ""
                            for i, info in enumerate(same_location_info):
                                if i > 0:
                                    popup_info += "<br>"
                                popup_info += f"<b>Абонент:</b> {info['абонент']}<br>"
                                popup_info += f"<b>Адреса БС:</b> {info['адреса']}<br>"
                                popup_info += f"<b>Азимут:</b> {info['азимут']}°"

                            # Маркер (додаємо в кінці, щоб був найвище над секторами)
                            folium.CircleMarker(
                                location=[lat, lon],
                                radius=8,
                                color='darkgreen',
                                fill=True,
                                fill_opacity=0.9,
                                weight=3,
                                zIndexOffset=1000,  # Високий z-index для відображення над секторами
                                popup=folium.Popup(popup_info, max_width=350, sticky=True)
                            ).add_to(grp)

                    except (ValueError, TypeError) as e:
                        logging.error(f"Помилка обробки координат для маркера {idx}: {e}")
                        continue

        def add_distances_to_map(map_obj, distances, groups):
            """Додає відстані на карту"""
            for m in distances:
                d = parse_datetime(m.get('Дата', ''))
                if not d or d not in groups:
                    continue

                grp = groups[d]

                try:
                    lat1 = float(m.get('Широта 1', 0))
                    lon1 = float(m.get('Довгота 1', 0))
                    lat2 = float(m.get('Широта 2', 0))
                    lon2 = float(m.get('Довгота 2', 0))
                    distance = m.get('Відстань (м)', 'Невідомо')

                    # Спочатку додаємо коло максимальної відстані (щоб було позаду)
                    max_dist = float(self.max_distance.get()) if hasattr(self, 'max_distance') else 1000
                    folium.Circle(
                        location=[lat1, lon1],
                        radius=max_dist,
                        color='blue', fill=True, fill_opacity=0.1,
                        popup=folium.Popup(
                            f"<b>Максимальна відстань:</b> {max_dist} м",
                            max_width=200,
                            sticky=True
                        )
                    ).add_to(grp)

                    # Потім лінію відстані
                    folium.PolyLine(
                        locations=[[lat1, lon1], [lat2, lon2]],
                        color='red', weight=3, opacity=0.9,
                        popup=folium.Popup(
                            f"<b>Відстань:</b> {distance} м",
                            max_width=200,
                            sticky=True
                        )
                    ).add_to(grp)

                    # В кінці маркери для точок (щоб були попереду)
                    for idx in ['1', '2']:
                        lat = float(m.get(f'Широта {idx}', 0))
                        lon = float(m.get(f'Довгота {idx}', 0))

                        # Отримуємо дані для підказки
                        abonent = m.get(f'Абонент А {idx}', 'Невідомо')
                        base_station_name = get_base_station_name(m, idx)

                        folium.CircleMarker(
                            location=[lat, lon],
                            radius=8,
                            color='red',
                            fill=True,
                            fill_opacity=0.9,
                            weight=3,
                            zIndexOffset=1000,  # Високий z-index
                            popup=folium.Popup(
                                f"<b>Абонент:</b> {abonent}<br>"
                                f"<b>Адреса БС:</b> {base_station_name}<br>"
                                f"<b>Відстань:</b> {distance} м",
                                max_width=300,
                                sticky=True
                            )
                        ).add_to(grp)

                except (ValueError, TypeError) as e:
                    logging.error(f"Помилка обробки відстані: {e}")

        # Основна логіка методу
        try:
            current_time = _get_current_datetime_and_user()

            # КРИТИЧНО: Логування початку
            logging.info(f"ПОЧАТОК create_meetings_map_with_excel: {current_time}")

            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END, f"\n{current_time}\n🗺️ Створення карт та файлів трафіку\n")
                self.log_text.see(tk.END)

            # Розділяємо зустрічі за типом
            sectors = [m for m in meetings if m.get('Тип збігу') == 'Сектори']
            distances = [m for m in meetings if m.get('Тип збігу') == 'Відстань']

            maps_created = 0

            # --- Карта секторів ---
            if sectors:
                # Збирання дат для секторів
                dates_sectors = sorted({
                    d for m in sectors
                    if (d := parse_datetime(m.get('Дата', '')))
                })

                if dates_sectors:
                    # Створюємо карту
                    try:
                        lats = [float(m.get('Широта 1', 0)) for m in sectors if m.get('Широта 1')]
                        lons = [float(m.get('Довгота 1', 0)) for m in sectors if m.get('Довгота 1')]

                        if lats and lons:
                            lat0 = np.mean(lats)
                            lon0 = np.mean(lons)
                        else:
                            lat0, lon0 = 50.4501, 30.5234  # Київ за замовчуванням

                    except (ValueError, TypeError):
                        lat0, lon0 = 50.4501, 30.5234  # Київ за замовчуванням

                    map_sectors = folium.Map(
                        location=[lat0, lon0],
                        zoom_start=12,
                        tiles='OpenStreetMap'
                    )

                    # Створюємо групи та додаємо елементи
                    groups_sectors = create_date_groups(map_sectors, dates_sectors)
                    add_sectors_to_map(map_sectors, sectors, groups_sectors)

                    # Додаємо контроль шарів (зліва вгорі)
                    folium.LayerControl(position='topleft', collapsed=False).add_to(map_sectors)

                    # Зберігаємо карту
                    sectors_filename = filename.replace('.html', '_sectors.html')
                    map_sectors.save(sectors_filename)
                    logging.info(f"Карта секторів збережена: {sectors_filename}")
                    maps_created += 1

            # --- Карта відстаней ---
            if distances:
                # Збирання дат для відстаней
                dates_distances = sorted({
                    d for m in distances
                    if (d := parse_datetime(m.get('Дата', '')))
                })

                if dates_distances:
                    # Створюємо карту
                    try:
                        lats = [float(m.get('Широта 1', 0)) for m in distances if m.get('Широта 1')]
                        lons = [float(m.get('Довгота 1', 0)) for m in distances if m.get('Довгота 1')]

                        if lats and lons:
                            lat0 = np.mean(lats)
                            lon0 = np.mean(lons)
                        else:
                            lat0, lon0 = 50.4501, 30.5234  # Київ за замовчуванням

                    except (ValueError, TypeError):
                        lat0, lon0 = 50.4501, 30.5234  # Київ за замовчуванням

                    map_distances = folium.Map(
                        location=[lat0, lon0],
                        zoom_start=12,
                        tiles='OpenStreetMap'
                    )

                    # Створюємо групи та додаємо елементи
                    groups_distances = create_date_groups(map_distances, dates_distances)
                    add_distances_to_map(map_distances, distances, groups_distances)

                    # Додаємо контроль шарів (зліва вгорі)
                    folium.LayerControl(position='topleft', collapsed=False).add_to(map_distances)

                    # Зберігаємо карту
                    distances_filename = filename.replace('.html', '_distances.html')
                    map_distances.save(distances_filename)
                    logging.info(f"Карта відстаней збережена: {distances_filename}")
                    maps_created += 1

            # =================================================================
            # КРИТИЧНО ВАЖЛИВО: ВИКЛИК create_excel_traffic_reports
            # =================================================================

            logging.info("🚀 ВИКЛИК create_excel_traffic_reports")

            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END, f"🚀 Виклик створення файлів трафіку...\n")
                self.log_text.see(tk.END)

            excel_success = False
            try:
                # ОБОВ'ЯЗКОВИЙ ВИКЛИК!
                excel_success = self.create_excel_traffic_reports(meetings, filename)
                logging.info(f"create_excel_traffic_reports результат: {excel_success}")
            except Exception as e:
                logging.error(f"Помилка create_excel_traffic_reports: {str(e)}")
                if hasattr(self, 'log_text'):
                    self.log_text.insert(tk.END, f"❌ Помилка файлів трафіку: {str(e)}\n")
                    self.log_text.see(tk.END)

            # Логування результатів
            if hasattr(self, 'log_text'):
                log_message = (
                    f"{current_time}\n"
                    f"Створено карти:\n"
                    f" - секторів: {len(sectors)} записів\n"
                    f" - відстаней: {len(distances)} записів\n"
                    f"Унікальних дат секторів: {len(dates_sectors) if 'dates_sectors' in locals() else 0}\n"
                    f"Унікальних дат відстаней: {len(dates_distances) if 'dates_distances' in locals() else 0}\n"
                    f"Файли карт збережено: {maps_created}\n"
                    f"Файли трафіку створено: {'✅' if excel_success else '❌'}\n"
                )

                if sectors and 'dates_sectors' in locals() and dates_sectors:
                    log_message += f" - {filename.replace('.html', '_sectors.html')}\n"
                if distances and 'dates_distances' in locals() and dates_distances:
                    log_message += f" - {filename.replace('.html', '_distances.html')}\n"

                self.log_text.insert(tk.END, log_message)
                self.log_text.see(tk.END)

            logging.info(f"ЗАВЕРШЕНО create_meetings_map_with_excel: карти={maps_created}, excel={excel_success}")

            return True

        except Exception as e:
            error_msg = f"Помилка створення карт: {str(e)}"
            logging.error(error_msg)

            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END, f"{_get_current_datetime_and_user()}\n{error_msg}\n")
                self.log_text.see(tk.END)

            return False

    def add_sector_to_map(self, feature_group, lat, lon, azimuth, radius, angle, color='red'):
        """
        Додає сектор до карти.

        Args:
            feature_group: FeatureGroup для додавання сектору
            lat: широта центру сектора
            lon: довгота центру сектора
            azimuth: азимут сектора в градусах
            radius: радіус сектора в метрах
            angle: кут сектора в градусах
            color: колір сектора (за замовчуванням 'red')

        Returns:
            folium.vector_layers.PolyLine: об'єкт сектору
        """
        import math

        # Конвертуємо азимут в математичний кут (проти годинникової стрілки від осі X)
        math_angle = (90 - azimuth) % 360

        # Розраховуємо початковий та кінцевий кути сектора
        start_angle = math_angle - angle / 2
        end_angle = math_angle + angle / 2

        # Створюємо точки для полігону сектора
        points = []
        points.append([lat, lon])  # Центр сектора

        # Додаємо точки по дузі
        num_points = 32  # кількість точок для апроксимації дуги
        for i in range(num_points + 1):
            current_angle = math.radians(start_angle + (end_angle - start_angle) * i / num_points)
            # Розраховуємо зміщення в метрах
            dx = radius * math.cos(current_angle)
            dy = radius * math.sin(current_angle)
            # Конвертуємо метри в градуси
            dlat = dy / 111111
            dlon = dx / (111111 * math.cos(math.radians(lat)))
            points.append([lat + dlat, lon + dlon])

        points.append([lat, lon])  # Замикаємо полігон

        # Створюємо сектор
        sector = folium.PolyLine(
            locations=points,
            color=color,
            weight=2,
            opacity=0.8,
            fill=True,
            fill_opacity=0.2
        )

        # Додаємо сектор до групи
        feature_group.add_child(sector)

        return sector

    def create_excel_traffic_reports(self, meetings, filename):
        """
        АЛЬТЕРНАТИВНИЙ підхід - використовує дані з основного Excel файлу
        Замість temp_db_path використовує meetings дані напряму
        """

        # Контроль рекурсії
        if not hasattr(self.__class__, '_excel_creation_lock'):
            self.__class__._excel_creation_lock = False

        if self.__class__._excel_creation_lock:
            return False

        self.__class__._excel_creation_lock = True

        def _get_current_datetime_and_user():
            return "2025-07-21 10:55:27 - McNeal1994"

        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            import os
            import logging
            import tkinter as tk
            import sqlite3
            import glob

            current_time = _get_current_datetime_and_user()

            # КРИТИЧНО: Логування початку роботи
            logging.info(f"ПОЧАТОК create_excel_traffic_reports: {current_time}")

            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END,
                                     f"\n{current_time}\n🔄 АЛЬТЕРНАТИВНИЙ підхід - пошук БД через основний Excel\n")
                self.log_text.see(tk.END)

            # ===== АЛЬТЕРНАТИВНИЙ ПОШУК БД =====

            db_path = None

            # Варіант 1: Через self.temp_db_path
            if hasattr(self, 'temp_db_path') and self.temp_db_path and os.path.exists(self.temp_db_path):
                db_path = self.temp_db_path
                if hasattr(self, 'log_text'):
                    self.log_text.insert(tk.END, f"✅ Використовуємо self.temp_db_path: {db_path}\n")
                    self.log_text.see(tk.END)

            # Варіант 2: Через основний Excel файл
            if not db_path:
                main_excel_file = filename.replace('.html', '.xlsx')

                if hasattr(self, 'log_text'):
                    self.log_text.insert(tk.END, f"🔍 Пошук через основний Excel: {main_excel_file}\n")
                    self.log_text.see(tk.END)

                if os.path.exists(main_excel_file):
                    # Читаємо основний Excel для отримання інформації про БД
                    try:
                        df_main = pd.read_excel(main_excel_file)

                        if hasattr(self, 'log_text'):
                            self.log_text.insert(tk.END, f"📊 Основний Excel: {len(df_main)} записів\n")
                            self.log_text.see(tk.END)

                        # Шукаємо .db файли в тій же директорії
                        excel_dir = os.path.dirname(main_excel_file)
                        db_pattern = os.path.join(excel_dir, "*.db")
                        found_dbs = glob.glob(db_pattern)

                        if hasattr(self, 'log_text'):
                            self.log_text.insert(tk.END, f"🔍 Знайдено .db файлів: {len(found_dbs)}\n")
                            for db_file in found_dbs:
                                file_size = os.path.getsize(db_file)
                                self.log_text.insert(tk.END, f"  {os.path.basename(db_file)}: {file_size} байт\n")
                            self.log_text.see(tk.END)

                        if found_dbs:
                            # Беремо найновіший файл
                            db_path = max(found_dbs, key=os.path.getmtime)
                            self.temp_db_path = db_path

                            if hasattr(self, 'log_text'):
                                self.log_text.insert(tk.END, f"✅ Вибрано БД: {os.path.basename(db_path)}\n")
                                self.log_text.see(tk.END)

                    except Exception as e:
                        if hasattr(self, 'log_text'):
                            self.log_text.insert(tk.END, f"❌ Помилка читання Excel: {str(e)}\n")
                            self.log_text.see(tk.END)

            # Варіант 3: Створення БД з meetings
            if not db_path:
                if hasattr(self, 'log_text'):
                    self.log_text.insert(tk.END, f"🔧 Створення БД з meetings даних\n")
                    self.log_text.see(tk.END)

                try:
                    # Створюємо тимчасову БД
                    import tempfile
                    temp_db = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
                    db_path = temp_db.name
                    temp_db.close()

                    # Підключаємося до БД
                    conn = sqlite3.connect(db_path)

                    # Створюємо таблицю з meetings
                    meetings_df = pd.DataFrame(meetings)
                    meetings_df.to_sql('meetings_table', conn, if_exists='replace', index=False)

                    conn.close()

                    self.temp_db_path = db_path

                    if hasattr(self, 'log_text'):
                        file_size = os.path.getsize(db_path)
                        self.log_text.insert(tk.END, f"✅ БД створена з meetings: {file_size} байт\n")
                        self.log_text.see(tk.END)

                except Exception as e:
                    if hasattr(self, 'log_text'):
                        self.log_text.insert(tk.END, f"❌ Помилка створення БД: {str(e)}\n")
                        self.log_text.see(tk.END)

            # Варіант 4: Створення файлів безпосередньо з meetings
            if not db_path:
                if hasattr(self, 'log_text'):
                    self.log_text.insert(tk.END, f"📋 Створення файлів безпосередньо з meetings\n")
                    self.log_text.see(tk.END)

                return create_excel_from_meetings_directly(meetings, filename, current_time, self)

            # Якщо БД знайдено, працюємо з нею
            if hasattr(self, 'log_text'):
                file_size = os.path.getsize(db_path)
                self.log_text.insert(tk.END, f"🎯 ВИКОРИСТОВУЄМО БД: {os.path.basename(db_path)} ({file_size} байт)\n")
                self.log_text.see(tk.END)

            # ===== ЗБІР ДАТ З MEETINGS =====

            dates_sectors = []
            dates_distances = []

            for meeting in meetings:
                if isinstance(meeting, dict):
                    meeting_type = meeting.get('Тип збігу', '').strip()
                    date = meeting.get('Дата', '').strip()

                    if date:
                        if meeting_type == 'Сектори':
                            dates_sectors.append(date)
                        elif meeting_type == 'Відстань':
                            dates_distances.append(date)

            dates_sectors = sorted(list(set(dates_sectors)))
            dates_distances = sorted(list(set(dates_distances)))

            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END, f"📅 Дати секторів: {dates_sectors}\n")
                self.log_text.insert(tk.END, f"📅 Дати відстаней: {dates_distances}\n")
                self.log_text.see(tk.END)

            # ===== ОТРИМАННЯ ДАНИХ З БД =====

            def get_db_data(target_dates):
                if not target_dates:
                    return pd.DataFrame()

                try:
                    conn = sqlite3.connect(db_path)
                    cursor = conn.cursor()

                    # Отримуємо таблиці
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                    tables = [row[0] for row in cursor.fetchall()]

                    all_data = []

                    for table_name in tables:
                        try:
                            # Простий запит
                            placeholders = ','.join(['?' for _ in target_dates])
                            query = f"SELECT * FROM [{table_name}] WHERE [Дата] IN ({placeholders})"

                            df = pd.read_sql_query(query, conn, params=target_dates)

                            if not df.empty:
                                all_data.append(df)
                                logging.info(f"Таблиця {table_name}: {len(df)} записів")

                        except Exception as e:
                            logging.error(f"Помилка таблиці {table_name}: {str(e)}")
                            continue

                    conn.close()

                    if all_data:
                        return pd.concat(all_data, ignore_index=True)
                    else:
                        return pd.DataFrame()

                except Exception as e:
                    logging.error(f"Помилка БД: {str(e)}")
                    return pd.DataFrame()

            # ===== СТВОРЕННЯ EXCEL ФАЙЛІВ =====

            def create_excel(data, output_path, title):
                if data.empty:
                    return False

                try:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Трафік"

                    columns = list(data.columns)

                    # Заголовок
                    ws['A1'] = f"ЗВЕДЕНИЙ ТРАФІК - {title.upper()}"
                    ws['A1'].font = Font(bold=True, size=14)

                    ws['A2'] = f"Створено: {current_time}"
                    ws['A2'].font = Font(size=10)

                    # Заголовки колонок
                    for col_idx, col_name in enumerate(columns, 1):
                        cell = ws.cell(row=4, column=col_idx, value=col_name)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                    # Дані
                    for row_idx, (_, data_row) in enumerate(data.iterrows(), 5):
                        for col_idx, col_name in enumerate(columns, 1):
                            ws.cell(row=row_idx, column=col_idx, value=str(data_row.get(col_name, '')))

                    wb.save(output_path)
                    wb.close()

                    return os.path.exists(output_path)

                except Exception as e:
                    logging.error(f"Помилка створення Excel: {str(e)}")
                    return False

            # ===== ОСНОВНА ОБРОБКА =====

            files_created = 0

            # Сектори
            if dates_sectors:
                sectors_data = get_db_data(dates_sectors)
                if not sectors_data.empty:
                    sectors_file = filename.replace('.html', '_sectors_traffic.xlsx')
                    if create_excel(sectors_data, sectors_file, 'Зустрічі по секторам'):
                        files_created += 1
                        logging.info(f"Створено файл секторів: {sectors_file}")

            # Відстані
            if dates_distances:
                distances_data = get_db_data(dates_distances)
                if not distances_data.empty:
                    distances_file = filename.replace('.html', '_distance_traffic.xlsx')
                    if create_excel(distances_data, distances_file, 'Зустрічі по відстані'):
                        files_created += 1
                        logging.info(f"Створено файл відстаней: {distances_file}")

            # Результат
            logging.info(f"ЗАВЕРШЕНО create_excel_traffic_reports: створено {files_created} файлів")

            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END, f"\n🏁 РЕЗУЛЬТАТ: {files_created} файлів створено\n")
                self.log_text.see(tk.END)

            self.__class__._excel_creation_lock = False
            return files_created > 0

        except Exception as e:
            self.__class__._excel_creation_lock = False

            error_msg = f"Критична помилка: {str(e)}"
            logging.error(error_msg)

            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END, f"\n❌ {error_msg}\n")
                self.log_text.see(tk.END)

            return False

    def create_excel_from_meetings_directly(meetings, filename, current_time, self_obj):
        """
        Створює Excel файли безпосередньо з meetings даних
        Використовується як запасний варіант
        """
        try:
            if hasattr(self_obj, 'log_text'):
                self_obj.log_text.insert(tk.END, f"📋 Створення з meetings: {len(meetings)} записів\n")
                self_obj.log_text.see(tk.END)

            # Розділяємо meetings по типах
            sectors_meetings = [m for m in meetings if m.get('Тип збігу') == 'Сектори']
            distances_meetings = [m for m in meetings if m.get('Тип збігу') == 'Відстань']

            files_created = 0

            # Функція створення Excel з meetings
            def create_excel_from_list(data_list, output_path, title):
                if not data_list:
                    return False

                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Дані"

                    # Отримуємо всі ключі
                    all_keys = set()
                    for item in data_list:
                        all_keys.update(item.keys())

                    columns = sorted(list(all_keys))

                    # Заголовок
                    ws['A1'] = f"ДАНІ ЗУСТРІЧЕЙ - {title.upper()}"
                    ws['A1'].font = Font(bold=True, size=14)

                    ws['A2'] = f"Створено: {current_time}"
                    ws['A2'].font = Font(size=10)

                    ws['A3'] = f"Записів: {len(data_list)}"
                    ws['A3'].font = Font(bold=True)

                    # Заголовки колонок
                    for col_idx, col_name in enumerate(columns, 1):
                        cell = ws.cell(row=5, column=col_idx, value=col_name)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                    # Дані
                    for row_idx, data_item in enumerate(data_list, 6):
                        for col_idx, col_name in enumerate(columns, 1):
                            ws.cell(row=row_idx, column=col_idx, value=str(data_item.get(col_name, '')))

                    wb.save(output_path)
                    wb.close()

                    return os.path.exists(output_path)

                except Exception as e:
                    logging.error(f"Помилка створення Excel з meetings: {str(e)}")
                    return False

            # Створюємо файли
            if sectors_meetings:
                sectors_file = filename.replace('.html', '_sectors_traffic.xlsx')
                if create_excel_from_list(sectors_meetings, sectors_file, 'Сектори'):
                    files_created += 1
                    logging.info(f"Створено файл секторів з meetings: {sectors_file}")

            if distances_meetings:
                distances_file = filename.replace('.html', '_distance_traffic.xlsx')
                if create_excel_from_list(distances_meetings, distances_file, 'Відстані'):
                    files_created += 1
                    logging.info(f"Створено файл відстаней з meetings: {distances_file}")

            if hasattr(self_obj, 'log_text'):
                self_obj.log_text.insert(tk.END, f"📋 Створено з meetings: {files_created} файлів\n")
                self_obj.log_text.see(tk.END)

            return files_created > 0

        except Exception as e:
            logging.error(f"Помилка create_excel_from_meetings_directly: {str(e)}")
            return False

    def create_merge_files_button(self):
        """
        Створює кнопку "Об'єднати файли" в інтерфейсі
        """
        import tkinter as tk
        from tkinter import ttk

        def _get_current_datetime_and_user():
            return "2025-07-21 11:07:14 - McNeal1994"

        # Створюємо фрейм для кнопки якщо не існує
        if not hasattr(self, 'merge_frame'):
            self.merge_frame = ttk.Frame(self.main_frame)
            self.merge_frame.pack(fill='x', padx=5, pady=5)

        # Кнопка об'єднання файлів
        self.merge_button = ttk.Button(
            self.merge_frame,
            text="📋 Об'єднати файли",
            command=self.merge_traffic_files,
            style='Accent.TButton'
        )
        self.merge_button.pack(side='left', padx=(0, 10))

        # Додаткова інформація
        self.merge_info_label = ttk.Label(
            self.merge_frame,
            text="Об'єднує всі файли трафіку в один зведений звіт",
            foreground='gray'
        )
        self.merge_info_label.pack(side='left', padx=(10, 0))

    def fix_merge_functionality():
        """
        ВИПРАВЛЕННЯ: Додає метод merge_traffic_files до правильного класу
        2025-07-21 11:16:23 - McNeal1994
        """
        import sys
        import logging

        def _get_current_datetime_and_user():
            return "2025-07-21 11:16:23 - McNeal1994"

        current_time = _get_current_datetime_and_user()
        logging.info(f"ВИПРАВЛЕННЯ merge_traffic_files: {current_time}")

        # Список можливих назв класів
        possible_class_names = [
            'DataProcessor',
            'TrafficTab',
            'MainApplication',
            'App',
            'TrafficAnalyzer'
        ]

        classes_found = []
        methods_added = 0

        # Шукаємо всі можливі класи
        for module_name, module in sys.modules.items():
            for class_name in possible_class_names:
                if hasattr(module, class_name):
                    target_class = getattr(module, class_name)
                    classes_found.append(f"{module_name}.{class_name}")

                    # Додаємо методи до класу
                    target_class.merge_traffic_files = merge_traffic_files
                    target_class.create_merge_files_button = create_merge_files_button

                    methods_added += 1
                    logging.info(f"Методи додано до {module_name}.{class_name}")

        print(f"🔧 {current_time}")
        print(f"Знайдено класів: {len(classes_found)}")
        print(f"Методи додано до {methods_added} класів")

        if classes_found:
            print("Класи з доданими методами:")
            for class_info in classes_found:
                print(f"  ✅ {class_info}")

        return methods_added > 0

    def merge_traffic_files(self):
        """
        ВИПРАВЛЕНИЙ метод об'єднання файлів трафіку
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        import os
        import glob
        import logging
        import tkinter as tk
        from tkinter import filedialog, messagebox
        from datetime import datetime

        def _get_current_datetime_and_user():
            return "2025-07-21 11:16:23 - McNeal1994"

        try:
            current_time = _get_current_datetime_and_user()
            logging.info(f"ПОЧАТОК merge_traffic_files: {current_time}")

            # Логування в UI якщо можливо
            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END, f"\n{current_time}\n🔗 ОБ'ЄДНАННЯ ФАЙЛІВ ТРАФІКУ\n")
                self.log_text.see(tk.END)

            # Визначаємо директорію для пошуку
            search_dir = "."

            # Пробуємо різні способи отримання поточної директорії
            if hasattr(self, 'current_directory'):
                search_dir = self.current_directory
            elif hasattr(self, 'last_directory'):
                search_dir = self.last_directory
            elif hasattr(self, 'working_directory'):
                search_dir = self.working_directory

            # Діалог вибору директорії
            selected_dir = filedialog.askdirectory(
                title="Виберіть директорію з файлами трафіку",
                initialdir=search_dir
            )

            if not selected_dir:
                if hasattr(self, 'log_text'):
                    self.log_text.insert(tk.END, "❌ Скасовано користувачем\n")
                    self.log_text.see(tk.END)
                return False

            # Пошук файлів трафіку
            traffic_patterns = [
                "*_sectors_traffic.xlsx",
                "*_distance_traffic.xlsx",
                "*traffic*.xlsx"
            ]

            found_files = []
            for pattern in traffic_patterns:
                pattern_path = os.path.join(selected_dir, pattern)
                files = glob.glob(pattern_path)
                found_files.extend(files)

            # Видаляємо дублікати
            found_files = sorted(list(set(found_files)))

            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END, f"🔍 Знайдено файлів: {len(found_files)}\n")
                for file_path in found_files:
                    self.log_text.insert(tk.END, f"  📄 {os.path.basename(file_path)}\n")
                self.log_text.see(tk.END)

            if not found_files:
                messagebox.showwarning(
                    "Файли не знайдено",
                    "У вибраній директорії не знайдено файлів трафіку.\n\n"
                    "Шукаємо файли з назвами:\n"
                    "- *_sectors_traffic.xlsx\n"
                    "- *_distance_traffic.xlsx\n"
                    "- *traffic*.xlsx"
                )
                return False

            # Читання та об'єднання файлів
            all_data = []
            file_stats = []

            for file_path in found_files:
                try:
                    if hasattr(self, 'log_text'):
                        self.log_text.insert(tk.END, f"📖 Читання: {os.path.basename(file_path)}\n")
                        self.log_text.see(tk.END)

                    df = pd.read_excel(file_path)

                    if not df.empty:
                        # Додаємо метадані
                        df['_source_file'] = os.path.basename(file_path)
                        df['_merge_time'] = current_time

                        all_data.append(df)
                        file_stats.append({
                            'file': os.path.basename(file_path),
                            'records': len(df),
                            'size': os.path.getsize(file_path)
                        })

                        logging.info(f"Прочитано {file_path}: {len(df)} записів")

                except Exception as e:
                    logging.error(f"Помилка читання {file_path}: {str(e)}")
                    if hasattr(self, 'log_text'):
                        self.log_text.insert(tk.END, f"  ❌ Помилка: {str(e)}\n")
                        self.log_text.see(tk.END)

            if not all_data:
                messagebox.showerror("Помилка", "Не вдалося прочитати жоден файл")
                return False

            # Об'єднання даних
            merged_data = pd.concat(all_data, ignore_index=True)

            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END, f"🔗 Об'єднано: {len(merged_data)} записів\n")
                self.log_text.see(tk.END)

            # Вибір файлу для збереження
            output_file = filedialog.asksaveasfilename(
                title="Збереження зведеного звіту",
                defaultextension=".xlsx",
                filetypes=[("Excel файли", "*.xlsx")],
                initialdir=selected_dir,
                initialname=f"merged_traffic_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

            if not output_file:
                return False

            # Створення Excel звіту
            wb = Workbook()
            wb.remove(wb.active)

            # Основний аркуш з даними
            ws = wb.create_sheet("Зведені дані")

            # Заголовок
            ws['A1'] = "ЗВЕДЕНИЙ ЗВІТ ТРАФІКУ"
            ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

            ws['A2'] = f"Створено: {current_time}"
            ws['A2'].font = Font(size=11)

            ws['A3'] = f"Файлів об'єднано: {len(found_files)}, Записів: {len(merged_data)}"
            ws['A3'].font = Font(bold=True)

            # Видаляємо службові колонки для відображення
            display_columns = [col for col in merged_data.columns if not col.startswith('_')]

            # Заголовки колонок
            for col_idx, col_name in enumerate(display_columns, 1):
                cell = ws.cell(row=5, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Дані
            for row_idx, (_, data_row) in enumerate(merged_data[display_columns].iterrows(), 6):
                for col_idx, col_name in enumerate(display_columns, 1):
                    ws.cell(row=row_idx, column=col_idx, value=str(data_row.get(col_name, '')))

            # Аркуш статистики
            ws_stats = wb.create_sheet("Статистика")

            ws_stats['A1'] = "СТАТИСТИКА ОБ'ЄДНАННЯ"
            ws_stats['A1'].font = Font(bold=True, size=14)

            ws_stats['A3'] = "Файл"
            ws_stats['B3'] = "Записів"
            ws_stats['C3'] = "Розмір (байт)"

            for row_idx, stats in enumerate(file_stats, 4):
                ws_stats.cell(row=row_idx, column=1, value=stats['file'])
                ws_stats.cell(row=row_idx, column=2, value=stats['records'])
                ws_stats.cell(row=row_idx, column=3, value=stats['size'])

            # Збереження
            wb.save(output_file)
            wb.close()

            # Підтвердження успіху
            if os.path.exists(output_file):
                file_size = os.path.getsize(output_file)
                success_msg = f"Зведений звіт створено!\n\nФайл: {os.path.basename(output_file)}\nОб'єднано файлів: {len(found_files)}\nЗаписів: {len(merged_data)}\nРозмір: {file_size} байт"

                logging.info(f"Зведений звіт створено: {output_file}")

                if hasattr(self, 'log_text'):
                    self.log_text.insert(tk.END, f"✅ УСПІХ: {os.path.basename(output_file)} ({file_size} байт)\n")
                    self.log_text.see(tk.END)

                messagebox.showinfo("Успіх!", success_msg)
                return True
            else:
                messagebox.showerror("Помилка", "Файл не було створено")
                return False

        except Exception as e:
            error_msg = f"Помилка об'єднання файлів: {str(e)}"
            logging.error(error_msg)

            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END, f"❌ {error_msg}\n")
                self.log_text.see(tk.END)

            messagebox.showerror("Помилка", error_msg)
            return False

    def create_merge_files_button(self):
        """
        Створює кнопку об'єднання файлів
        """
        import tkinter as tk
        from tkinter import ttk

        # Знаходимо або створюємо фрейм для кнопок
        if not hasattr(self, 'merge_frame'):
            # Пробуємо різні можливі батьківські контейнери
            parent = None
            if hasattr(self, 'main_frame'):
                parent = self.main_frame
            elif hasattr(self, 'content_frame'):
                parent = self.content_frame
            elif hasattr(self, 'frame'):
                parent = self.frame
            elif hasattr(self, 'master'):
                parent = self.master

            if parent:
                self.merge_frame = ttk.Frame(parent)
                self.merge_frame.pack(fill='x', padx=5, pady=5)
            else:
                print("❌ Не вдалося знайти батьківський контейнер для кнопки")
                return False

        # Створюємо кнопку
        self.merge_button = ttk.Button(
            self.merge_frame,
            text="🔗 Об'єднати файли",
            command=self.merge_traffic_files
        )
        self.merge_button.pack(side='left', padx=(0, 10))

        # Інформаційний лейбл
        info_label = ttk.Label(
            self.merge_frame,
            text="Об'єднує всі файли трафіку в один зведений звіт",
            foreground='gray'
        )
        info_label.pack(side='left', padx=(10, 0))

        print("✅ Кнопка 'Об'єднати файли' додана до інтерфейсу")
        return True

    # Запускаємо виправлення
    if __name__ == "__main__":
        fix_merge_functionality()

    def filter_traffic_by_datetime(self, traffic_files, filter_file, time_window_minutes, progress_bar, root,
                                   output_dir):
        """Фільтрація трафіку за датою і часом."""
        # Використовуємо той самий код, але як метод класу
        return self.data_processor.filter_traffic_by_datetime(
            traffic_files,
            filter_file,
            time_window_minutes,
            progress_bar,
            root,
            output_dir
        )

    @staticmethod
    def assign_district_by_coords(lat, lon):
        """Присвоєння району за координатами."""
        if pd.isna(lat) or pd.isna(lon):
            return "Неизвестный"
        lat_zone = int(lat * 10) / 10
        lon_zone = int(lon * 10) / 10
        return f"Зона_{lat_zone}_{lon_zone}"

    @staticmethod
    def get_daily_trajectories(df):
        """Отримання щоденних траєкторій."""
        trajectories = {}
        grouped = df.groupby('Дата')
        for date, group in grouped:
            group = group.sort_values('Час')
            districts = group['district'].tolist()
            cleaned_districts = [districts[0]] if districts else []
            for d in districts[1:]:
                if d != cleaned_districts[-1]:
                    cleaned_districts.append(d)
            if cleaned_districts:
                trajectories[date] = cleaned_districts
        return trajectories

    @staticmethod
    def is_subsequence(sub, full, min_match_length):
        """Перевірка чи є послідовність підпослідовністю."""
        if len(sub) < min_match_length:
            return False
        pos = 0
        for item in full:
            if pos < len(sub) and item == sub[pos]:
                pos += 1
        return pos >= min_match_length and pos == len(sub)

    @staticmethod
    def parse_time(time_str):
        """Парсинг часу з різних форматів."""
        if isinstance(time_str, time):
            return time_str
        if pd.isna(time_str) or not str(time_str).strip():
            logging.warning(f"Пустое время: {time_str}")
            return None
        formats = ['%H:%M:%S', '%H:%M', '%I:%M:%S %p', '%I:%M %p']
        for fmt in formats:
            try:
                return pd.to_datetime(str(time_str), format=fmt, errors='coerce').time()
            except ValueError:
                continue
        logging.warning(f"Некорректный формат времени: {time_str}")
        return None

